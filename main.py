from time import sleep
from playwright.sync_api import sync_playwright
import os
import json
import urllib.request
from openpyxl import load_workbook


args=[
     '--disable-blink-features=AutomationControlled',
     '--start-maximized',
     '--disable-infobars',
     '--no-sandbox',
     '--disable-dev-shm-usage',
     '--disable-extensions',
     '--remote-debugging-port=0',
     '--disable-web-security',
     '--enable-features=WebRTCPeerConnectionWithBlockIceAddresses',
     '--force-webrtc-ip-handling-policy=disable_non_proxied_udp',
 ]

chrome_path = os.path.join(os.getcwd(), "chrome-win/chrome.exe")
# Define a function that opens the browser and returns the browser and contex
storage_state_file = os.path.join(os.getcwd(), "storage_state.json")
def is_storage_state_valid(file_path):
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
                return bool(data)  # true false depend data exist
        except json.JSONDecodeError as e:
            print(f"JSON decode error: {e}")
            return False
    else:
        open(file_path, 'w').close()  # Create an empty file
        return False

def cookie_save():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            executable_path=str(chrome_path),
            headless=False,
            args=args,
        )
        if is_storage_state_valid(storage_state_file):
            context = browser.new_context(storage_state=storage_state_file, no_viewport=True)
        else:
            context = browser.new_context(no_viewport=True)

        page = context.new_page()
        page.goto("https://www.facebook.com/")

        input("Press Enter to save data: ")

        # Save the storage state (including cookies)
        context.storage_state(path=storage_state_file)
        print("Data saved...")
        browser.close()



if not is_storage_state_valid(storage_state_file):
    cookie_save()

# //img[@alt='No photo description available.']


workbook = load_workbook('data.xlsx')
sheet = workbook['Sheet1']

# Iterate over rows and print values
i = 0
with sync_playwright() as playwright:
    browser = playwright.chromium.launch(
        executable_path=str(chrome_path),
        headless=False,
        args=args,
    )
    if is_storage_state_valid(storage_state_file):
        context = browser.new_context(storage_state=storage_state_file, no_viewport=True)
        page = context.new_page()
        for row in sheet.iter_rows(values_only=True):
            if i!=0:
                try:
                    page.goto(row[0])
                    img = page.locator("(//img[@alt='No photo description available.'])[1]")
                    if img.count() > 0:
                        print(f"src : {row[0]} ")
                        print(f" {str(i)} Founded")
                        image_src = img.get_attribute('src', timeout=600000)
                        print(image_src)
                        if not os.path.exists("images"):
                            os.mkdir("images")
                        urllib.request.urlretrieve(image_src, f"images/{row[1]}.png")
                except:
                    pass
            i+=1
    else:
        print("Login Failed")

browser.close()